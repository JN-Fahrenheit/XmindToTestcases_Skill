import os
import json
from datetime import datetime
import re
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from xmindparser import xmind_to_dict

# --- 配置 ---
# 获取当前脚本所在的目录
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
Xmind_DIR = os.path.join(BASE_DIR, 'assert')
TEMPLATE_PATH = os.path.join(BASE_DIR, 'templates', '测试用例模板.xlsx')
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')

# --- 辅助函数 ---

def find_xmind_file(directory):
    """在指定目录中查找.xmind文件"""
    for filename in os.listdir(directory):
        if filename.endswith('.xmind'):
            return os.path.join(directory, filename), filename
    return None, None

def parse_details(text):
    """从五级节点文本中解析前置条件、步骤和预期结果"""
    if not isinstance(text, str):
        return "", [], []

    precondition = ""
    steps = []
    expectations = []

    pre_match = re.search(r'前提：(.*?)(?=(步骤：|预期：|$))', text, re.DOTALL)
    if pre_match:
        precondition = pre_match.group(1).strip()

    steps_match = re.search(r'步骤：(.*?)(?=(前提：|预期：|$))', text, re.DOTALL)
    if steps_match:
        steps_text = steps_match.group(1).strip()
        steps = [s.strip() for s in re.split(r'\n|\d+\.', steps_text) if s.strip()]

    exp_match = re.search(r'预期：(.*?)(?=(前提：|步骤：|$))', text, re.DOTALL)
    if exp_match:
        exp_text = exp_match.group(1).strip()
        expectations = [e.strip() for e in re.split(r'\n|\d+\.', exp_text) if e.strip()]

    return precondition, steps, expectations


def process_node(node, path, test_cases):
    """递归遍历节点，提取测试用例"""
    if not isinstance(node, dict) or 'title' not in node:
        return

    current_path = path + [node['title']]
    
    # 当达到第四级节点时，处理第五级节点
    if len(current_path) == 4:
        case_name = current_path[3]
        if 'topics' in node and node['topics']:
            for sub_node in node['topics']:
                if len(current_path) == 4: # 确保是四级节点的子节点
                    details_text = sub_node.get('title', '')
                    precondition, steps, expectations = parse_details(details_text)
                    
                    # 配对步骤和预期
                    max_len = max(len(steps), len(expectations))
                    if max_len == 0: # 如果没有步骤和预期，也生成一行
                         test_cases.append({
                            "用例名称": case_name,
                            "所属模块": current_path[0] if len(current_path) > 0 else "",
                            "前置条件": precondition,
                            "步骤描述": "",
                            "预期结果": ""
                        })
                    else:
                        for i in range(max_len):
                            step = steps[i] if i < len(steps) else ""
                            expectation = expectations[i] if i < len(expectations) else ""
                            test_cases.append({
                                "用例名称": case_name,
                                "所属模块": current_path[0] if len(current_path) > 0 else "",
                                "前置条件": precondition,
                                "步骤描述": step,
                                "预期结果": expectation
                            })
        else:
            # 第四级节点没有子节点
            print(f"警告: 测试场景 '{case_name}' 没有详细的测试用例（第五级节点）。")
            test_cases.append({
                "用例名称": case_name,
                "所属模块": current_path[0] if len(current_path) > 0 else "",
                "前置条件": "N/A",
                "步骤描述": "N/A",
                "预期结果": "N/A"
            })


    # 递归处理子节点
    if 'topics' in node and node['topics']:
        for sub_node in node['topics']:
            process_node(sub_node, current_path, test_cases)


def write_to_excel(test_cases, template_path, output_path):
    """将测试用例写入Excel文件，并处理单元格合并"""
    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except FileNotFoundError:
        print(f"错误: 模板文件未找到 '{template_path}'")
        return

    # 获取表头
    headers = [cell.value for cell in ws[1]]
    
    # 数据从第二行开始
    row_idx = 2
    
    # 按用例名称分组
    from itertools import groupby
    
    for case_name, group in groupby(test_cases, key=lambda x: x['用例名称']):
        group_list = list(group)
        num_rows = len(group_list)
        
        start_row = row_idx
        
        for i, case in enumerate(group_list):
            # 填充数据
            for col_idx, header in enumerate(headers, 1):
                if header in case:
                    ws.cell(row=row_idx + i, column=col_idx, value=case[header])

        # 合并单元格
        if num_rows > 1:
            # 合并用例名称
            if "用例名称" in headers:
                col = headers.index("用例名称") + 1
                ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + num_rows - 1, end_column=col)
            
            # 合并前置条件
            if "前置条件" in headers:
                col = headers.index("前置条件") + 1
                ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + num_rows - 1, end_column=col)

        row_idx += num_rows

    wb.save(output_path)
    print(f"成功生成测试用例文件: {output_path}")


def main(target_xmind=None):
    """主函数"""
    # 1. 检查并找到xmind文件
    if target_xmind:
        xmind_file_path = os.path.join(Xmind_DIR, target_xmind)
        xmind_filename = target_xmind
        if not os.path.exists(xmind_file_path):
            print(f"错误: 指定的文件 '{target_xmind}' 不在 '{Xmind_DIR}' 目录中。")
            return
    else:
        xmind_file_path, xmind_filename = find_xmind_file(Xmind_DIR)
        if not xmind_file_path:
            print(f"错误: 在 '{Xmind_DIR}' 目录中未找到.xmind文件。")
            return

    print(f"找到Xmind文件: {xmind_filename}")

    # 2. 检查模板文件是否存在
    if not os.path.exists(TEMPLATE_PATH):
        print(f"错误: 测试用例模板文件不存在于 '{TEMPLATE_PATH}'。")
        return

    # 3. 解析Xmind文件
    try:
        xmind_data = xmind_to_dict(xmind_file_path)
    except Exception as e:
        print(f"错误: 解析Xmind文件失败: {e}")
        return
    
    if not xmind_data:
        print("错误: Xmind文件内容为空或格式不正确。")
        return

    # 4. 遍历所有画布（Sheet）及其下的所有一级节点并提取测试用例
    all_test_cases = []
    
    # 常见的通用根节点名称，如果根节点是这些名称，则跳过根节点，将其子节点作为一级节点
    GENERIC_ROOT_NAMES = ["测试用例", "测试用例集", "测试计划", "中央主题", "XMind", " ", ""]

    for sheet in xmind_data:
        root_topic = sheet.get('topic')
        if not root_topic:
            continue
            
        root_title = root_topic.get('title', '').strip()
        
        # 判断根节点是否为“一级节点”
        # 如果根节点名称太通用，或者是空的，且它有子节点，则将其子节点视为一级节点
        if (root_title in GENERIC_ROOT_NAMES or not root_title) and 'topics' in root_topic and root_topic['topics']:
            for level1_node in root_topic['topics']:
                process_node(level1_node, [], all_test_cases)
        else:
            # 否则，根节点本身就是第一个一级节点
            # 注意：这里我们还要检查 root_topic 是否有“浮动主题”（Floating Topics）
            # 在某些 XMind 版本中，浮动主题也可能被放在 root_topic 的 topics 列表中，
            # 但它们在逻辑上与 root_topic 是并列的一级节点。
            # 为了简单起见，如果根节点有标题，我们从根节点开始。
            process_node(root_topic, [], all_test_cases)

    if not all_test_cases:
        print("未能在Xmind文件中找到任何有效的测试用例。请检查文件结构是否符合约定。")
        return

    # 5. 生成输出文件名并写入Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_filename = os.path.splitext(xmind_filename)[0]
    output_filename = f"{base_filename}_测试用例_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, output_filename)

    write_to_excel(all_test_cases, TEMPLATE_PATH, output_path)


if __name__ == "__main__":
    import sys
    # 确保输出目录存在
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    
    target = sys.argv[1] if len(sys.argv) > 1 else None
    main(target)
